import importlib
import os
import sqlite3
import tempfile
import threading
import time
import unittest
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import text


def _reload_modules_for_sqlite(sqlite_path):
    os.environ['SQLITE_PATH'] = sqlite_path

    import database
    import db_utils
    import parser
    import supplier_service

    importlib.reload(database)
    importlib.reload(db_utils)
    importlib.reload(supplier_service)
    importlib.reload(parser)

    return database, db_utils, parser, supplier_service


def _build_logs_xlsx(path, sku='SKU-1'):
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value='Отчет по товарам')

    ws.cell(row=2, column=1, value='meta')
    ws.cell(row=2, column=2, value='meta')
    ws.cell(row=2, column=3, value='meta')
    ws.cell(row=2, column=4, value='01.04.2026')
    ws.cell(row=2, column=5, value='01.04.2026')

    ws.cell(row=3, column=1, value='Наименование')
    ws.cell(row=3, column=2, value='Остаток на начало периода')
    ws.cell(row=3, column=3, value='Остаток на конец периода')
    ws.cell(row=3, column=4, value='Прих')
    ws.cell(row=3, column=5, value='Расх')

    ws.cell(row=4, column=1, value=sku)
    ws.cell(row=4, column=2, value=10)
    ws.cell(row=4, column=3, value=9)
    ws.cell(row=4, column=4, value=2)
    ws.cell(row=4, column=5, value=3)

    wb.save(path)


def _build_spoils_xlsx(path, sku='SKU-1'):
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value='Списания')
    headers = ['Дата', 'B', 'C', 'D', 'Товар', 'Количество', 'G', 'Причина']
    for idx, value in enumerate(headers, start=1):
        ws.cell(row=2, column=idx, value=value)

    ws.cell(row=3, column=1, value=datetime(2026, 4, 1))
    ws.cell(row=3, column=5, value=sku)
    ws.cell(row=3, column=6, value=1)
    ws.cell(row=3, column=8, value='Повреждение')

    wb.save(path)


def _build_price_xlsx(path, sku='SKU-1', supplier='Supplier A'):
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value='Прайс-лист')
    ws.cell(row=2, column=1, value='')

    # Row 3 is the first data row (raw.iloc[2:]).
    ws.cell(row=3, column=2, value=sku)          # B
    ws.cell(row=3, column=5, value=100)          # E sale_price
    ws.cell(row=3, column=6, value=50)           # F purchase_price
    ws.cell(row=3, column=7, value=0)            # G discount
    ws.cell(row=3, column=8, value='6')          # H packaging
    ws.cell(row=3, column=10, value=supplier)    # J supplier_name

    wb.save(path)


class IntegrationWorkflowTests(unittest.TestCase):
    def setUp(self):
        self._old_sqlite_path = os.environ.get('SQLITE_PATH')
        self._paths_to_cleanup = []

    def _new_temp_path(self, suffix):
        path = str(Path(tempfile.gettempdir()) / f'colizeum_integration_{uuid.uuid4().hex}{suffix}')
        self._paths_to_cleanup.append(path)
        return path

    def tearDown(self):
        try:
            import database
            database.engine.dispose()
        except Exception:
            pass

        for path in self._paths_to_cleanup:
            try:
                if os.path.exists(path):
                    os.remove(path)
            except OSError:
                pass

        if self._old_sqlite_path is None:
            os.environ.pop('SQLITE_PATH', None)
        else:
            os.environ['SQLITE_PATH'] = self._old_sqlite_path

        # Rebind modules back to default environment after each test.
        import database
        import db_utils
        import parser
        import supplier_service

        importlib.reload(database)
        importlib.reload(db_utils)
        importlib.reload(supplier_service)
        importlib.reload(parser)

    def test_startup_init_db_migrates_legacy_uploaded_files_schema(self):
        db_path = self._new_temp_path('.db')

        conn = sqlite3.connect(db_path)
        try:
            conn.execute(
                """
                CREATE TABLE uploaded_files (
                    id INTEGER PRIMARY KEY,
                    filename TEXT UNIQUE NOT NULL,
                    upload_date DATETIME
                )
                """
            )
            conn.execute(
                "INSERT INTO uploaded_files (filename, upload_date) VALUES (?, ?)",
                ('spoils::legacy_file.xlsx', '2026-04-01 00:00:00')
            )
            conn.commit()
        finally:
            conn.close()

        database, _, _, _ = _reload_modules_for_sqlite(db_path)
        database.init_db()

        conn = sqlite3.connect(db_path)
        try:
            cols = {
                row[1]
                for row in conn.execute('PRAGMA table_info(uploaded_files)').fetchall()
            }
            self.assertIn('date_from', cols)
            self.assertIn('date_to', cols)
            self.assertIn('file_type', cols)

            row = conn.execute(
                "SELECT file_type FROM uploaded_files WHERE filename='spoils::legacy_file.xlsx'"
            ).fetchone()
            self.assertIsNotNone(row)
            self.assertEqual(row[0], 'spoils')
        finally:
            conn.close()

    def test_upload_workflow_logs_spoils_price_updates_business_tables(self):
        db_path = self._new_temp_path('.db')
        database, db_utils, parser, _ = _reload_modules_for_sqlite(db_path)
        database.init_db()

        with tempfile.TemporaryDirectory() as tmpdir:
            logs_path = str(Path(tmpdir) / 'logs.xlsx')
            spoils_path = str(Path(tmpdir) / 'spoils.xlsx')
            price_path = str(Path(tmpdir) / 'price.xlsx')

            _build_logs_xlsx(logs_path, sku='SKU-1')
            _build_spoils_xlsx(spoils_path, sku='SKU-1')
            _build_price_xlsx(price_path, sku='SKU-1', supplier='Supplier A')

            with open(logs_path, 'rb') as f:
                parser.parse_and_save_file(f)
            with open(spoils_path, 'rb') as f:
                parser.parse_and_save_spoils_file(f)
            with open(price_path, 'rb') as f:
                parser.parse_and_save_price_list_file(f)

        session = db_utils.get_session()
        try:
            product_count = session.query(database.Product).count()
            sale_count = session.query(database.Sale).count()
            spoil_count = session.query(database.Spoil).count()
            net_sales = session.query(database.NetSale).all()
            uploaded_types = {f.file_type for f in session.query(database.UploadedFile).all()}

            suppliers = session.query(database.Supplier).all()
            supplier = next((s for s in suppliers if s.name == 'Supplier A'), None)
            self.assertIsNotNone(supplier)

            price_item_count = session.query(database.PriceListItem).count()

            self.assertEqual(product_count, 1)
            self.assertEqual(sale_count, 1)
            self.assertEqual(spoil_count, 1)
            self.assertEqual(len(net_sales), 1)
            self.assertAlmostEqual(float(net_sales[0].quantity), 2.0, places=6)
            self.assertSetEqual(uploaded_types, {'logs', 'spoils', 'price'})

            self.assertEqual(float(supplier.delivery_cost), 1000.0)
            self.assertEqual(str(supplier.delivery_time), '3')
            self.assertEqual(float(supplier.min_order), 5000.0)

            self.assertEqual(price_item_count, 1)
        finally:
            session.close()

    def test_concurrent_update_parameters_retries_after_sqlite_lock(self):
        db_path = self._new_temp_path('.db')
        database, db_utils, _, _ = _reload_modules_for_sqlite(db_path)
        database.init_db()

        lock_session = db_utils.get_session()
        lock_session.execute(text('BEGIN IMMEDIATE'))

        errors = []

        def _worker():
            try:
                db_utils.update_parameters({'quote_multiplicator': 2.5})
            except Exception as exc:  # pragma: no cover - assertion below handles failures
                errors.append(exc)

        thread = threading.Thread(target=_worker, daemon=True)
        thread.start()

        # Keep write lock long enough to force at least one retry.
        time.sleep(0.6)
        lock_session.rollback()
        lock_session.close()

        thread.join(timeout=6)

        self.assertFalse(thread.is_alive(), 'update_parameters did not finish after lock release')
        self.assertEqual(errors, [])

        verify_session = db_utils.get_session()
        try:
            param = verify_session.query(database.Parameter).filter_by(key='quote_multiplicator').first()
            self.assertIsNotNone(param)
            self.assertAlmostEqual(float(param.value), 2.5, places=6)
        finally:
            verify_session.close()


if __name__ == '__main__':
    unittest.main()
